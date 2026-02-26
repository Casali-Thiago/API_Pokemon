import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

Main_Url = "https://pokeapi.co/api/v2"


def lista_de_pokemon(limit: int = 20) -> list[dict]:
    url = f"{Main_Url}/pokemon?limit={limit}"
    response = requests.get(url, timeout=10)
    response.raise_for_status()
    return response.json()["results"]


def acharPokemon(url: str) -> dict:
    response = requests.get(url, timeout=10)
    response.raise_for_status()
    data = response.json()

    types = ", ".join(t["type"]["name"] for t in data["types"])

    return {
        "id": data["id"],
        "nome": data["name"].capitalize(),
        "altura": data["height"],
        "peso": data["weight"],
        "tipos": types,
        "base_experience": data.get("base_experience", "N/A"),
    }


def create_excel(pokemons: list[dict], filename: str = "pokemon.xlsx") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pokémons"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="E3350D")
    center = Alignment(horizontal="center", vertical="center")
    border_side = Side(style="thin", color="CCCCCC")
    cell_border = Border(
        left=border_side, right=border_side, top=border_side, bottom=border_side
    )

    headers = ["ID", "Nome", "Altura (dm)", "Peso (hg)", "Tipos", "Base Experience"]
    col_widths = [8, 18, 14, 12, 22, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    alt_fill = PatternFill("solid", start_color="FFF3F3")

    for row_idx, pokemon in enumerate(pokemons, start=2):
        values = [
            pokemon["id"],
            pokemon["nome"],
            pokemon["altura"],
            pokemon["peso"],
            pokemon["tipos"],
            pokemon["base_experience"],
        ]
        fill = alt_fill if row_idx % 2 == 0 else None

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = center
            cell.border = cell_border
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 20

    wb.save(filename)
    print(f"Arquivo '{filename}'Pokémons Capturados!")


def main():
    print("Buscando lista de Pokémons...")
    pokemon_list = lista_de_pokemon(limit=20)

    pokemons = []
    for i, entry in enumerate(pokemon_list, start=1):
        print(f"[{i}/20] {entry['name'].capitalize()} encontrado")
        try:
            data = acharPokemon(entry["url"])
            pokemons.append(data)
        except requests.RequestException as e:
            print(f"  Erro ao buscar {entry['name']}: {e}")

    create_excel(pokemons)


if __name__ == "__main__":
    main()